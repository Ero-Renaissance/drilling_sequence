"""The import template must be correct BY CONSTRUCTION: whatever it teaches,
the importer accepts. The strongest check is the round trip — download the
template, upload it unchanged, and assert it imports cleanly (activities,
contracts AND the per-project Readiness sheet) with exactly the expected
(intentional) cross-terrain notice."""
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook


async def _project(client: AsyncClient) -> str:
    return (await client.post("/api/projects", json={"name": "Template"})).json()["id"]


@pytest.mark.asyncio
async def test_template_downloads_with_schedule_readiness_and_guidance(
    client: AsyncClient,
) -> None:
    pid = await _project(client)
    r = await client.get(f"/api/projects/{pid}/activities/import-template")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames[:2] == ["Schedule", "Readiness"]
    assert "Guidance" in wb.sheetnames

    ws = wb["Schedule"]
    header = [c.value for c in ws[1]]
    assert header[:3] == ["Location", "Rig Name", "HWU Name"]
    # Readiness is per PROJECT now — the Schedule sheet carries no gate columns.
    assert "Readiness Check" not in header
    assert "Readiness Check Status" not in header

    rd = wb["Readiness"]
    assert [c.value for c in rd[1]] == [
        "Project", "FDP", "LLI", "LOC", "FE", "FID", "EIA", "BUD",
    ]

    # The guidance sheet carries the canonical activity types (dropdown source).
    gd = wb["Guidance"]
    col_d = [gd.cell(row=i, column=4).value for i in range(2, 20)]
    assert "Well Cleanup/Test" in col_d
    assert "Rig Mobilisation and Intake" in col_d


@pytest.mark.asyncio
async def test_template_round_trips_through_the_importer(client: AsyncClient) -> None:
    pid = await _project(client)
    r = await client.get(f"/api/projects/{pid}/activities/import-template")
    assert r.status_code == 200

    up = await client.post(
        f"/api/projects/{pid}/activities/import?replace=true",
        files={
            "file": (
                "schedule-import-template.xlsx",
                io.BytesIO(r.content),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert up.status_code == 200, up.text
    body = up.json()
    assert body["imported"] == 3  # Well-1 (LAND rig), Well-2 (SWAMP twin), Well-3 (HWU)
    assert body["skipped"] == 0

    # Exactly ONE warning: the DELIBERATE cross-terrain twin ("Rig 1" on LAND and
    # SWAMP) that the template uses to demonstrate the identity convention. Any
    # other warning means the template teaches something the importer rejects.
    assert len(body["warnings"]) == 1, body["warnings"]
    assert "Rig 1" in body["warnings"][0] and "multiple terrains" in body["warnings"][0]

    # The terrain twins each land their OWN contract.
    contracts = (await client.get(f"/api/projects/{pid}/contracts")).json()
    pairs = {(c["rig_name"], c["terrain"], c["contract_end"]) for c in contracts}
    assert ("Rig 1", "LAND", "2030-12-31") in pairs
    assert ("Rig 1", "SWAMP", "2031-06-30") in pairs
    hwu = (await client.get(f"/api/projects/{pid}/hwu-contracts")).json()
    assert [(c["hwu_name"], c["contract_end"]) for c in hwu] == [("HWU 1", "2031-06-30")]

    # The Readiness sheet's per-PROJECT gates (incl. Behind and N/A) all imported.
    readiness = (await client.get(f"/api/projects/{pid}/readiness")).json()
    by_project = {row["well_project"]: row for row in readiness}
    assert set(by_project) == {"Project Alpha", "Project Beta"}
    alpha = by_project["Project Alpha"]
    assert alpha["activity_count"] == 2  # Well-1 + Well-2 share the project
    assert alpha["checks"]["FDP"]["status"] == "Completed"
    assert alpha["checks"]["LOC"]["status"] == "Behind"
    assert alpha["checks"]["EIA"]["status"] == "N/A"
    beta = by_project["Project Beta"]
    assert all(c["status"] == "On Track" for c in beta["checks"].values())


@pytest.mark.asyncio
async def test_planner_wording_behind_schedule_is_mapped_not_dropped(
    client: AsyncClient,
) -> None:
    """'Behind Schedule' — the planner's habitual wording — maps to Behind on the
    Readiness sheet instead of being dropped with a warning."""
    from openpyxl import Workbook

    pid = await _project(client)
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(["Location", "Rig Name", "HWU Name", "Activity Type", "Plan Type",
               "Project", "Well Name", "Start Date", "End Date",
               "Rig Contract Expiry Date", "HWU Contract Expiry Date", "Risk", "Comment"])
    ws.append(["LAND", "RIG_1", None, "Oil Development", "In Plan (Firm)", "PX",
               "W-1", "05/01/2026", "15/03/2026", None, None, "No Flood Risk", None])
    rd = wb.create_sheet("Readiness")
    rd.append(["Project", "FDP", "LLI", "LOC", "FE", "FID", "EIA", "BUD"])
    rd.append(["PX", "On Track", "Behind Schedule", "On Track", "On Track",
               "On Track", "On Track", "On Track"])
    buf = io.BytesIO()
    wb.save(buf)

    up = await client.post(
        f"/api/projects/{pid}/activities/import?replace=true",
        files={"file": ("s.xlsx", io.BytesIO(buf.getvalue()),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert up.status_code == 200, up.text
    assert up.json()["warnings"] == []

    readiness = (await client.get(f"/api/projects/{pid}/readiness")).json()
    assert readiness[0]["checks"]["LLI"]["status"] == "Behind"
