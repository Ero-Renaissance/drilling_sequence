"""Rig fleet optimization — engine and API (docs/rig-optimization-spec.md).

The engine tests anchor on the spec's worked example: a 6-well project on one
rig chains as 2w/2w/4w/2w/2w gaps → 540 days (≈17.8 months).
"""

import io
from datetime import date

import pytest
from httpx import AsyncClient

from app.services.rig_optimizer import (
    Assumptions,
    Options,
    optimize,
    optimize_terrain,
)

A = Assumptions()  # spec defaults: 76d well, 14d gap, batch 3, 28d, 45d move
STRICT = Options()  # finished in-year, no slip, no drill-ahead


@pytest.fixture(autouse=True)
def _default_heuristic_engine(monkeypatch):
    """Pin the engine to heuristic so these tests don't depend on whatever
    OPTIMIZER_ENGINE the local .env sets; milp tests opt in by re-patching."""
    from app.config import settings

    monkeypatch.setattr(settings, "optimizer_engine", "heuristic")


def _single_project(terrain: str, wells_by_year: dict[int, int]) -> dict:
    return {"terrain": terrain, "project": "P1", "wells_by_year": wells_by_year}


# ---------------------------------------------------------------------------
# The worked example (spec §3.2)
# ---------------------------------------------------------------------------

def test_six_well_chain_matches_spec_worked_example() -> None:
    """6 wells on ONE rig: gaps 2w,2w,4w,2w,2w → 6×76 + 84 = 540 days.

    Spread over two years (3+3) so a single rig is actually sufficient and the
    full chain is observable.
    """
    result = optimize_terrain("Land", {"P1": {2027: 3, 2028: 3}}, A, STRICT)
    assert result.feasible and result.rig_count == 1
    wells = result.rigs[0].wells
    assert [w.gap_kind for w in wells] == [
        "none", "inter_well", "inter_well", "batch", "inter_well", "inter_well",
    ]
    assert [w.gap_before_days for w in wells] == [0, 14, 14, 28, 14, 14]
    # End-to-end chain: 456 drilling + 84 gap days = 540 (the 4-week batch gap
    # REPLACES that slot's 2 weeks; it is not additive).
    span = (wells[-1].end - wells[0].start).days
    # The batch gap slot straddles the year boundary; the rig waits for 1 Jan
    # 2028 (drill-ahead off), so the span is >= the pure chain length.
    assert span >= 540
    # With drill-ahead allowed the chain is exactly back-to-back: 540 days.
    relaxed = optimize_terrain(
        "Land", {"P1": {2027: 3, 2028: 3}}, A, Options(allow_drill_ahead=True)
    )
    chain = relaxed.rigs[0].wells
    assert (chain[-1].end - chain[0].start).days == 540


def test_batch_counter_continues_across_years_by_default() -> None:
    """Wells 3→4 straddle the year boundary; the 4-week batch gap still applies
    (no reset on 1 January unless the toggle is on)."""
    result = optimize_terrain("Land", {"P1": {2027: 3, 2028: 3}}, A, STRICT)
    gaps = [w.gap_kind for w in result.rigs[0].wells]
    assert gaps[3] == "batch"

    reset = optimize_terrain(
        "Land", {"P1": {2027: 3, 2028: 3}}, A, Options(batch_reset_on_new_year=True)
    )
    gaps = [w.gap_kind for w in reset.rigs[0].wells]
    assert gaps[3] == "inter_well"  # counter reset on New Year → plain 2-week move


def test_project_move_is_terrain_specific() -> None:
    """Between projects: land rigs move in 45 days; swamp and offshore in 30."""
    land = optimize_terrain("Land", {"P1": {2027: 1}, "P2": {2027: 1}}, A, STRICT)
    assert land.feasible and land.rig_count == 1
    assert land.rigs[0].wells[1].gap_kind == "project_move"
    assert land.rigs[0].wells[1].gap_before_days == 45

    for terrain in ("Swamp", "SWO"):
        r = optimize_terrain(terrain, {"P1": {2027: 1}, "P2": {2027: 1}}, A, STRICT)
        assert r.feasible and r.rig_count == 1
        assert r.rigs[0].wells[1].gap_kind == "project_move"
        assert r.rigs[0].wells[1].gap_before_days == 30


# ---------------------------------------------------------------------------
# Fleet sizing + yearly constraint
# ---------------------------------------------------------------------------

def test_five_wells_one_year_needs_two_rigs() -> None:
    """5 wells in one year on one rig = 5×76 + 4 gaps > 365 days → 2 rigs
    (concurrency at a project is allowed, spec §9.1)."""
    result = optimize_terrain("Swamp", {"P1": {2027: 5}}, A, STRICT)
    assert result.feasible
    assert result.rig_count == 2
    assert result.binding == {"project": "P1", "year": 2027}


def test_four_wells_one_year_fits_one_rig() -> None:
    """4×76 + (14+14+28) = 360 ≤ 365 → a single rig just delivers it."""
    result = optimize_terrain("Land", {"P1": {2027: 4}}, A, STRICT)
    assert result.feasible and result.rig_count == 1


def test_terrains_are_sealed_fleets() -> None:
    """Identical demand in two terrains → two independent 1-rig answers, never a
    shared rig."""
    results = optimize(
        [
            _single_project("Land", {2027: 2}),
            {"terrain": "Swamp", "project": "P9", "wells_by_year": {2027: 2}},
        ],
        A,
        STRICT,
    )
    assert [r.terrain for r in results] == ["Land", "Swamp"]
    assert all(r.rig_count == 1 for r in results)


def test_spudded_delivery_is_more_lenient_than_finished() -> None:
    """A 5-well chain of 70-day wells starts its last well on day 350: too late
    to FINISH in-year (needs a 2nd rig), but fine to SPUD in-year (1 rig)."""
    short = Assumptions(well_duration_days=70)
    finished = optimize_terrain("Land", {"P1": {2027: 5}}, short, STRICT)
    assert finished.rig_count == 2

    spudded = optimize_terrain(
        "Land", {"P1": {2027: 5}}, short, Options(delivery="spudded")
    )
    assert spudded.rig_count == 1
    last = spudded.rigs[0].wells[-1]
    assert last.start <= date(2027, 12, 31) < last.end  # spudded in-year, finishes later


def test_slip_allowance_relaxes_the_deadline() -> None:
    """One rig finishes the 5th default well on day 450 — a 90-day grace past
    31 December makes the single-rig chain acceptable."""
    strict = optimize_terrain("Land", {"P1": {2027: 5}}, A, STRICT)
    assert strict.rig_count == 2
    slack = optimize_terrain(
        "Land", {"P1": {2027: 5}}, A, Options(allow_slip_days=90)
    )
    assert slack.rig_count == 1


def test_drill_ahead_uses_idle_year() -> None:
    """1 well in 2027 + 10 in 2028: strict crams all ten inside 2028 (3 rigs);
    drill-ahead lets rigs start 2028's wells during idle 2027 (fewer rigs)."""
    demand = {"P1": {2027: 1, 2028: 10}}
    strict = optimize_terrain("Land", demand, A, STRICT)
    ahead = optimize_terrain("Land", demand, A, Options(allow_drill_ahead=True))
    assert ahead.rig_count < strict.rig_count


def test_structural_infeasibility_is_reported_not_hidden() -> None:
    """A well longer than the year window can never finish in-year: the result
    says infeasible and names the project/year."""
    long_wells = Assumptions(well_duration_days=400)
    result = optimize_terrain("Land", {"P1": {2027: 1}}, long_wells, STRICT)
    assert not result.feasible
    assert result.infeasible_wells == [{"project": "P1", "year": 2027}]


def test_empty_demand_needs_zero_rigs() -> None:
    result = optimize_terrain("Land", {"P1": {}}, A, STRICT)
    assert result.feasible and result.rig_count == 0


def test_fleet_profile_and_utilization_shape() -> None:
    result = optimize_terrain("Land", {"P1": {2027: 2, 2028: 2}}, A, STRICT)
    assert set(result.rigs_active_per_year) == {2027, 2028}
    assert all(0 < u <= 1 for u in result.utilization_per_rig.values())


# ---------------------------------------------------------------------------
# API — authorization and round trips
# ---------------------------------------------------------------------------

_REQUEST = {
    "demand": [
        {"terrain": "Land", "project": "Alpha", "wells_by_year": {"2027": 2}},
        {"terrain": "Swamp", "project": "Beta", "wells_by_year": {"2027": 5}},
    ]
}


@pytest.mark.asyncio
async def test_optimize_endpoint_planner_allowed(client: AsyncClient) -> None:
    r = await client.post("/api/optimizer/rig-fleet", json=_REQUEST)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["engine"] == "heuristic"
    by_terrain = {t["terrain"]: t for t in body["results"]}
    assert by_terrain["Land"]["rig_count"] == 1
    assert by_terrain["Swamp"]["rig_count"] == 2


@pytest.mark.asyncio
async def test_optimize_endpoint_denied_without_grant(noplan_client: AsyncClient) -> None:
    r = await noplan_client.post("/api/optimizer/rig-fleet", json=_REQUEST)
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_milp_falls_back_on_unsupported_relaxation(
    client: AsyncClient, monkeypatch
) -> None:
    """The exact engine models only the strict in-year policy; a relaxation that
    couples years (drill-ahead here) degrades to the heuristic with a warning.
    Deterministic regardless of whether OR-Tools is installed."""
    from app.config import settings

    monkeypatch.setattr(settings, "optimizer_engine", "milp")
    req = {**_REQUEST, "options": {"allow_drill_ahead": True}}
    r = await client.post("/api/optimizer/rig-fleet", json=req)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["engine"] == "heuristic"
    assert body["warning"] and "drill-ahead" in body["warning"]


@pytest.mark.asyncio
async def test_milp_engine_used_end_to_end(client: AsyncClient, monkeypatch) -> None:
    """With OR-Tools installed and a strict-policy request, the API reports the
    milp engine and returns the exact (lower) fleet — 5 rigs where the heuristic
    would say 6."""
    pytest.importorskip("ortools")
    from app.config import settings

    monkeypatch.setattr(settings, "optimizer_engine", "milp")
    req = {
        "demand": [
            {"terrain": "Swamp", "project": "S1",
             "wells_by_year": {"2027": 5, "2028": 5, "2029": 14, "2030": 6}},
            {"terrain": "Swamp", "project": "S2",
             "wells_by_year": {"2029": 4, "2030": 7, "2031": 2}},
        ]
    }
    r = await client.post("/api/optimizer/rig-fleet", json=req)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["engine"] == "milp"
    assert body["warning"] is None
    assert body["results"][0]["rig_count"] == 5


def test_milp_engine_matches_or_beats_heuristic() -> None:
    """Engine invariant: the exact fleet is never larger than the heuristic's,
    and strictly smaller where the greedy over-assigns. Both stay feasible."""
    ortools = pytest.importorskip("ortools")  # noqa: F841 — skip if solver absent
    from app.services.rig_optimizer import Assumptions, Options, optimize, run

    cases = [
        [{"terrain": "Land", "project": "P", "wells_by_year": {"2027": 5}}],
        [
            {
                "terrain": "Swamp",
                "project": "S1",
                "wells_by_year": {"2027": 5, "2028": 5, "2029": 14, "2030": 6},
            },
            {
                "terrain": "Swamp",
                "project": "S2",
                "wells_by_year": {"2029": 4, "2030": 7, "2031": 2},
            },
        ],
    ]
    a_, opts_ = Assumptions(), Options()
    beat_at_least_once = False
    for demand in cases:
        h = optimize(demand, a_, opts_)
        m, engine, warning = run(demand, a_, opts_, "milp")
        assert engine == "milp" and warning is None
        for rh, rm in zip(h, m):
            assert rm.feasible == rh.feasible
            assert rm.rig_count <= rh.rig_count
            if rm.rig_count < rh.rig_count:
                beat_at_least_once = True
            # Materialized schedule is physically sound: in-year, no overlap.
            for rig in rm.rigs:
                for w in rig.wells:
                    assert date(w.year, 1, 1) <= w.start
                    assert w.end <= date(w.year, 12, 31)
                ws = sorted(rig.wells, key=lambda x: x.start)
                for a, b in zip(ws, ws[1:]):
                    assert b.start >= a.end
    assert beat_at_least_once, "expected the exact engine to beat greedy somewhere"


def test_milp_well_labels_unique_when_project_splits_across_rigs() -> None:
    """A project's wells that split across rigs (concurrency) must each carry a
    distinct label — the per-rig well counter must not restart at 'Well 1' on
    every rig (regression: two bars both read 'Well 1')."""
    pytest.importorskip("ortools")
    from app.services.rig_optimizer import Assumptions, Options
    from app.services.rig_optimizer_milp import optimize_milp

    # 5 wells of one project in one year force a split across ≥2 rigs.
    demand = [
        {"terrain": "Land", "project": "Project Land 1", "wells_by_year": {"2027": 5}},
        {"terrain": "Land", "project": "Project Land 2", "wells_by_year": {"2027": 2}},
    ]
    result = optimize_milp(demand, Assumptions(), Options())[0]
    assert result.rig_count >= 2  # the 5-well project can't fit one rig in a year

    labels = [w.label for rig in result.rigs for w in rig.wells]
    assert len(labels) == len(set(labels)), f"duplicate well labels: {labels}"
    # And the split project's wells are numbered 1..5 exactly once each.
    p1 = sorted(
        w.label for rig in result.rigs for w in rig.wells if w.project == "Project Land 1"
    )
    assert p1 == [f"Project Land 1 · 2027 · Well {n}" for n in range(1, 6)]


@pytest.mark.asyncio
async def test_optimize_endpoint_validates_terrain(client: AsyncClient) -> None:
    bad = {"demand": [{"terrain": "Desert", "project": "X", "wells_by_year": {"2027": 1}}]}
    r = await client.post("/api/optimizer/rig-fleet", json=bad)
    assert r.status_code == 422


@pytest.mark.asyncio
async def test_parse_schedule_csv_roundtrip(client: AsyncClient) -> None:
    csv = (
        "Terrain,Project,2027,2028\n"
        "Land,Project 1,2,\n"
        "Swamp,Project 10,1,3\n"
        "Mars,Project X,1,\n"  # unknown terrain → reported, not silently dropped
    )
    r = await client.post(
        "/api/optimizer/parse-schedule",
        files={"file": ("schedule.csv", io.BytesIO(csv.encode()), "text/csv")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["years"] == [2027, 2028]
    assert len(body["demand"]) == 2
    assert body["demand"][1]["wells_by_year"] == {"2027": 1, "2028": 3}
    assert any("Mars" in issue for issue in body["issues"])


@pytest.mark.asyncio
async def test_parse_schedule_denied_without_grant(noplan_client: AsyncClient) -> None:
    r = await noplan_client.post(
        "/api/optimizer/parse-schedule",
        files={"file": ("s.csv", io.BytesIO(b"Terrain,Project,2027\n"), "text/csv")},
    )
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_parse_schedule_rejects_columnless_file(client: AsyncClient) -> None:
    r = await client.post(
        "/api/optimizer/parse-schedule",
        files={"file": ("s.csv", io.BytesIO(b"a,b\n1,2\n"), "text/csv")},
    )
    assert r.status_code == 422


@pytest.mark.asyncio
async def test_export_returns_workbook_with_expected_sheets(client: AsyncClient) -> None:
    r = await client.post("/api/optimizer/rig-fleet/export", json=_REQUEST)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "rig-optimization.xlsx" in r.headers["content-disposition"]

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Summary", "Rig Schedule", "Demand"]
    summary = wb["Summary"]
    assert summary["A1"].value == "Terrain"
    terrains = {summary.cell(row=i, column=1).value for i in (2, 3)}
    assert terrains == {"Land", "Swamp"}
    sched = wb["Rig Schedule"]
    # 2 Land + 5 Swamp wells scheduled → 7 data rows.
    assert sched.max_row == 1 + 7
    assert sched["B2"].value.endswith("Rig 1")


@pytest.mark.asyncio
async def test_export_denied_without_grant(noplan_client: AsyncClient) -> None:
    r = await noplan_client.post("/api/optimizer/rig-fleet/export", json=_REQUEST)
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_create_campaign_from_result(client: AsyncClient, db) -> None:
    """The optimizer→campaign bridge: a NEW Draft campaign with activities from
    the scheduled wells, rigs registered as PLANNED slots, provenance in the
    key notes, and a governance audit entry."""
    payload = {
        "name": "Q3 2027 — optimized draft",
        "default_activity_type": "Oil Development",
        "engine": "heuristic",
        "results": [
            {
                "terrain": "Land",
                "feasible": True,
                "rig_count": 1,
                "rigs": [
                    {
                        "name": "LAND Opt Rig 1",
                        "wells": [
                            {"project": "P-Alpha", "year": 2027, "label": "P-Alpha · WELL_1",
                             "start": "2027-01-10", "end": "2027-03-27", "gap_before_days": 0, "gap_kind": "none"},
                            {"project": "P-Alpha", "year": 2027, "label": "P-Alpha · WELL_2",
                             "start": "2027-04-10", "end": "2027-06-25", "gap_before_days": 14, "gap_kind": "inter_well"},
                        ],
                    }
                ],
                "rigs_active_per_year": {"2027": 1},
                "utilization_per_rig": {"LAND Opt Rig 1": 0.8},
            }
        ],
    }
    r = await client.post("/api/optimizer/create-campaign", json=payload)
    assert r.status_code == 201, r.text
    pid = r.json()["id"]

    acts = (await client.get(f"/api/projects/{pid}/activities")).json()
    assert len(acts) == 2
    assert {a["well_name"] for a in acts} == {"WELL_1", "WELL_2"}
    assert all(a["rig_name"] == "LAND Opt Rig 1" for a in acts)
    assert all(a["well_project"] == "P-Alpha" for a in acts)
    assert all(a["location"] == "LAND" for a in acts)
    assert all(a["activity_type"] == "Oil Development" for a in acts)

    # The optimizer's hypothetical rig is a PLANNED slot in the registry.
    units = (await client.get(f"/api/projects/{pid}/resources")).json()
    slot = next(u for u in units if u["name"] == "LAND Opt Rig 1")
    assert slot["is_placeholder"] is True

    # Provenance: key notes + governance audit entry.
    detail = (await client.get(f"/api/projects/{pid}")).json()
    assert "rig optimization" in detail["key_notes"]["body"]
    audit = (await client.get(f"/api/projects/{pid}/audit")).json()
    assert any("from rig optimization" in (e.get("new_value") or "") for e in audit)


@pytest.mark.asyncio
async def test_create_campaign_requires_planner_grant(noplan_client: AsyncClient) -> None:
    r = await noplan_client.post(
        "/api/optimizer/create-campaign",
        json={"name": "X", "default_activity_type": "Oil Development", "results": []},
    )
    # Empty results 422s at the schema; use a minimal valid body for the 403.
    assert r.status_code in (403, 422)
    r = await noplan_client.post(
        "/api/optimizer/create-campaign",
        json={
            "name": "X", "default_activity_type": "Oil Development",
            "results": [{"terrain": "Land", "feasible": True, "rig_count": 1,
                         "rigs": [{"name": "R", "wells": [{"project": "P", "year": 2027, "label": "P · W",
                                   "start": "2027-01-01", "end": "2027-02-01", "gap_before_days": 0, "gap_kind": "none"}]}],
                         "rigs_active_per_year": {}, "utilization_per_rig": {}}],
        },
    )
    assert r.status_code == 403, r.text


# ---------------------------------------------------------------------------
# Per-year completion cutoff (each year's last well must FINISH by its month)
# ---------------------------------------------------------------------------


def _all_ends(result) -> list:
    return [w.end for rig in result.rigs for w in rig.wells]


def test_completion_cutoff_bounds_every_finish_heuristic() -> None:
    """With a June cutoff on 2027, every 2027 well FINISHES by 30 June — and the
    squeezed window needs more rigs than the free year (the point: realism)."""
    a_cut = Assumptions(last_completion_month_by_year={2027: 6})
    free = optimize_terrain("Land", {"P1": {2027: 4}}, A, STRICT)
    cut = optimize_terrain("Land", {"P1": {2027: 4}}, a_cut, STRICT)
    assert free.feasible and cut.feasible
    assert all(e.month <= 6 for e in _all_ends(cut))
    assert cut.rig_count >= free.rig_count


def test_completion_cutoff_is_per_year() -> None:
    """Only the listed year is squeezed: 2027 ends by June, 2028 runs free."""
    a_cut = Assumptions(last_completion_month_by_year={2027: 6})
    r = optimize_terrain("Land", {"P1": {2027: 2, 2028: 4}}, a_cut, STRICT)
    assert r.feasible
    for rig in r.rigs:
        for w in rig.wells:
            if w.year == 2027:
                assert w.end.month <= 6
    # 2028 wells may finish after June (the free year keeps full capacity).
    assert any(w.year == 2028 and w.end.month > 6 for rig in r.rigs for w in rig.wells)


def test_completion_cutoff_is_hard_against_slip() -> None:
    """allow_slip_days relaxes the ordinary 31-December deadline, never an
    explicit cutoff — a flood window doesn't negotiate."""
    a_cut = Assumptions(last_completion_month_by_year={2027: 6})
    slippy = Options(allow_slip_days=60)
    r = optimize_terrain("Land", {"P1": {2027: 4}}, a_cut, slippy)
    assert r.feasible
    assert all(e.month <= 6 for e in _all_ends(r) )


def test_completion_cutoff_milp_matches() -> None:
    pytest.importorskip("ortools")
    from app.services.rig_optimizer_milp import optimize_milp

    a_cut = Assumptions(last_completion_month_by_year={2027: 6})
    f_tr = optimize_milp([_single_project("Land", {2027: 4})], A, STRICT)[0]
    c_tr = optimize_milp([_single_project("Land", {2027: 4})], a_cut, STRICT)[0]
    assert f_tr.feasible and c_tr.feasible
    assert all(w.end.month <= 6 for rig in c_tr.rigs for w in rig.wells)
    assert c_tr.rig_count >= f_tr.rig_count
    # The exact engine never needs more rigs than the greedy heuristic.
    heur = optimize_terrain("Land", {"P1": {2027: 4}}, a_cut, STRICT)
    assert c_tr.rig_count <= heur.rig_count


def test_completion_cutoff_infeasibility_is_reported() -> None:
    """A window shorter than one well's drilling time can't finish anything —
    reported as infeasible, never silently dropped."""
    a_hard = Assumptions(last_completion_month_by_year={2027: 2})  # 59 days < 76
    r = optimize_terrain("Land", {"P1": {2027: 1}}, a_hard, STRICT)
    assert not r.feasible
    assert r.infeasible_wells
