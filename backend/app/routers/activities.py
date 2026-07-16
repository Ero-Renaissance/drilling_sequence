import io
import json
import uuid
from datetime import datetime, timezone
from typing import Annotated

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import ValidationError
from sqlalchemy import delete, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth import get_current_user
from app.core.locks import assert_project_not_locked, ensure_activity_unlocked
from app.core.rbac import assert_member
from app.database import get_db
from app.models.activity import Activity
from app.models.activity_type_alias import ActivityTypeAlias
from app.models.audit import AuditLog
from app.models.hwu_contract import HwuContract
from app.models.project import Project, ProjectRole
from app.models.rig_contract import RigContract
from app.models.user import User
from app.schemas.activity import (
    ActivityCreate,
    ActivityCreateStrict,
    ActivityResponse,
    ActivityUpdate,
    ImportResponse,
)
from app.schemas.audit import AuditEntryResponse
from app.services.audit import ENTITY_VOCABULARY, governance_event
from app.services.data_processor import (
    CANONICAL_ACTIVITY_TYPES,
    SCHEDULE_SHEET,
    cross_terrain_resource_warnings,
    csv_df_to_db_rows,
    is_long_schedule,
    normalize_activity_type_key,
    parse_long_schedule,
    resolve_activity_type,
    unknown_activity_type_warnings,
    validate_csv_columns,
)
from app.services.registry import (
    ensure_activity_resources_registered,
    ensure_registered,
    lane_contract_rows,
)

router = APIRouter(prefix="/api/projects/{project_id}/activities", tags=["activities"])

# Import upload bounds. Real campaign sheets are a few hundred KB / a few
# thousand rows (7 rows per well); the caps only exist to stop a huge or
# decompression-bomb upload from exhausting memory. Mirrors the optimizer's
# _MAX_UPLOAD_BYTES pattern (app/routers/optimizer.py).
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000

CurrentUser = Annotated[User, Depends(get_current_user)]
DB = Annotated[AsyncSession, Depends(get_db)]

# Fields excluded from the audit log and conflict check
_AUDIT_EXCLUDE = {"expected_updated_at"}


def _normalize_ts(ts: datetime) -> datetime:
    """Ensure timezone-aware UTC datetime for comparison."""
    if ts.tzinfo is None:
        return ts.replace(tzinfo=timezone.utc)
    return ts.astimezone(timezone.utc)


@router.get("", response_model=list[ActivityResponse])
async def list_activities(
    project_id: uuid.UUID, current_user: CurrentUser, db: DB
) -> list[ActivityResponse]:
    # Reads are org-wide: any authenticated user may view campaign data.
    result = await db.execute(
        select(Activity)
        .where(Activity.project_id == project_id)
        .order_by(Activity.start_date)
    )
    return [ActivityResponse.model_validate(a) for a in result.scalars().all()]


# Terrain row order on every Gantt: Land → Swamp → Offshore (unknown sorts last).
_TERRAIN_ORDER = {"LAND": 0, "SWAMP": 1, "OFFSHORE": 2}


@router.get("/export")
async def export_activities(project_id: uuid.UUID, current_user: CurrentUser, db: DB):
    """Export the full campaign plan (every activity) as an Excel workbook —
    the rig sequence as a table, sorted to mirror the chart (terrain → resource
    → start). Read-only, so any authenticated user may download it. Readiness
    checks are intentionally excluded."""
    from fastapi.responses import StreamingResponse

    project = await db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    rows = (
        await db.execute(select(Activity).where(Activity.project_id == project_id))
    ).scalars().all()

    def _sort_key(a: Activity) -> tuple:
        resource = a.hwu_name or a.rig_name or ""
        return (_TERRAIN_ORDER.get((a.location or "").upper(), 99), resource, a.start_date)

    rows = sorted(rows, key=_sort_key)

    from openpyxl import Workbook  # already a vetted dependency (Excel import)
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Rig Sequence"
    headers = [
        "Location", "Resource Type", "Resource Name", "Well", "Project",
        "Activity Type", "Plan Type", "Risk", "Start", "End",
        "Duration (days)", "Comment", "Completed",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    date_fmt = "DD-MM-YYYY"
    for a in rows:
        resource_type = "HWU" if a.hwu_name else ("Rig" if a.rig_name else "")
        duration = (a.end_date - a.start_date).days if a.start_date and a.end_date else None
        completed = a.completed_at.date() if a.completed_at else None
        ws.append(
            [
                a.location, resource_type, a.hwu_name or a.rig_name, a.well_name,
                a.well_project, a.activity_type, a.plan_type, a.risk,
                a.start_date, a.end_date, duration, a.comment, completed,
            ]
        )
        r = ws.max_row
        ws.cell(row=r, column=9).number_format = date_fmt   # Start
        ws.cell(row=r, column=10).number_format = date_fmt  # End
        if completed is not None:
            ws.cell(row=r, column=13).number_format = date_fmt

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = "".join(c for c in project.name if c.isalnum() or c in " -_").strip() or "campaign"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe} - rig sequence.xlsx"'},
    )


@router.get("/import-template")
async def download_import_template(project_id: uuid.UUID, current_user: CurrentUser):
    """A blank, self-documenting Excel template for the schedule import.

    Static content (no project data), so any authenticated user may download it.
    Sheet 1 "Schedule": ONE ROW PER ACTIVITY (Location + Rig Name identify a
    physical rig — a LAND and a SWAMP "Rig 1" are two rigs), with Excel dropdowns
    on every enum column. Sheet 2 "Readiness": ONE ROW PER PROJECT — the seven
    sanction gates (FDP…BUD) are field-development-project attributes, not
    per-activity. Sheet 3 "Guidance": the rules + canonical vocabularies.
    """
    from datetime import date as _date

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    # ── Sheet 1: Schedule (one row per activity) ──────────────────────────────
    header = [
        "Location", "Rig Name", "HWU Name", "Activity Type", "Plan Type", "Project",
        "Well Name", "Start Date", "End Date", "Rig Contract Expiry Date",
        "HWU Contract Expiry Date", "Risk", "Comment",
    ]
    samples = [
        ["LAND", "Rig 1", None, "Gas Development", "In Plan (Firm)", "Project Alpha",
         "Well-1", _date(2026, 1, 15), _date(2026, 6, 30), _date(2030, 12, 31), None,
         "No Flood Risk", "One row per activity — set the project's readiness gates in the app"],
        ["SWAMP", "Rig 1", None, "Oil Development", "In Plan (Option)", "Project Alpha",
         "Well-2", _date(2026, 3, 1), _date(2026, 8, 31), _date(2031, 6, 30), None,
         "Flood Risk", "Same name as the LAND rig = a DIFFERENT physical rig (see Guidance)"],
        ["SWAMP", None, "HWU 1", "Well Repair/Safety", "In Plan (Firm)", "Project Beta",
         "Well-3", _date(2026, 9, 1), _date(2026, 11, 30), None, _date(2031, 6, 30),
         "No Flood Risk", "A row uses a rig OR an HWU, never both"],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = SCHEDULE_SHEET
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in samples:
        ws.append(row)
    for r in range(2, ws.max_row + 1):
        for c in (8, 9, 10, 11):  # Start, End, both expiry columns
            ws.cell(row=r, column=c).number_format = "DD/MM/YYYY"
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 42)

    # ── Sheet 2: Guidance — the rules + the canonical vocabularies ────────────
    gd = wb.create_sheet("Guidance")
    gd.append(["Rule", "Guidance"])
    for cell in gd[1]:
        cell.font = Font(bold=True)
    rules = [
        ("One row per activity", "The Schedule tab is ONE ROW PER ACTIVITY. Legacy files that "
         "repeated a well once per readiness gate still import — the repeats collapse."),
        ("Readiness is set in the app", "Sanction gates (FDP, LLI, LOC, FE, FID, EIA, BUD) are "
         "per FIELD-DEVELOPMENT PROJECT and are managed on the app's Readiness tab after "
         "import — they are NOT part of this upload, and re-importing never resets them."),
        ("Rig identity", "A rig is identified by Location + Rig Name. The same name in two "
         "locations is TWO physical rigs (e.g. a land 10K and a swamp 10K barge) — the import "
         "confirms this with an informational notice. Two rigs of the same class in the SAME "
         "location must be numbered (10K Rig 1, 10K Rig 2)."),
        ("Planned rigs", "Don't know the awarded rig yet? Use a class-style slot name "
         "(e.g. '10K Rig 3'). It registers as a PLANNED unit; rename it from the Fleet page "
         "when the contract is awarded — its schedule and contract follow automatically."),
        ("HWUs", "HWUs are mobile units: the same HWU Name anywhere is ONE unit, whatever "
         "the location."),
        ("Dates", "Day-first — DD/MM/YYYY or DD-MM-YYYY (e.g. 31/07/2026). Real Excel date "
         "cells work too. A month-first or impossible date rejects the whole upload."),
        ("Activity Type", "Use the exact canonical names in column D — anything else imports "
         "but charts in neutral grey until an admin adds it to the catalogue."),
        ("Rig Contract Expiry Date", "ONE date per physical rig. Terrain twins may each carry "
         "their own date. Leave blank when there is no contract yet — correct for planned "
         "units; the dashboard then raises a procurement alert."),
        ("Plan Type", "In Plan (Firm), In Plan (Option) or Out of Plan."),
        ("Risk", "Flood Risk or No Flood Risk."),
    ]
    for rule in rules:
        gd.append(list(rule))
    gd.column_dimensions["A"].width = 26
    gd.column_dimensions["B"].width = 110
    for r in range(2, gd.max_row + 1):
        gd.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Canonical activity types — listed for the planner AND feeding the dropdown.
    gd.cell(row=1, column=4, value="Canonical Activity Types").font = Font(bold=True)
    types = sorted(CANONICAL_ACTIVITY_TYPES)
    for i, t in enumerate(types, start=2):
        gd.cell(row=i, column=4, value=t)
    gd.column_dimensions["D"].width = 34
    wb.defined_names["ActivityTypes"] = DefinedName(
        "ActivityTypes", attr_text=f"Guidance!$D$2:$D${1 + len(types)}"
    )

    # ── Dropdowns on the Schedule sheet's enum columns ─────────────────────────
    def dropdown(formula: str, cells: str) -> None:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(cells)

    last = 2000  # generous editing room
    dropdown('"LAND,SWAMP,OFFSHORE"', f"A2:A{last}")
    dropdown("ActivityTypes", f"D2:D{last}")
    dropdown('"In Plan (Firm),In Plan (Option),Out of Plan"', f"E2:E{last}")
    dropdown('"Flood Risk,No Flood Risk"', f"L2:L{last}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="schedule-import-template.xlsx"'},
    )


@router.post("", response_model=ActivityResponse, status_code=status.HTTP_201_CREATED)
async def create_activity(
    project_id: uuid.UUID, payload: ActivityCreateStrict, current_user: CurrentUser, db: DB
) -> ActivityResponse:
    await assert_member(project_id, current_user, db, allowed_roles={ProjectRole.planner})
    # Adding an activity mutates the plan, so it is barred while a revision is
    # awaiting approval — the same gate as import (which also creates activities),
    # consistent with the update/complete/delete locks on existing activities.
    await assert_project_not_locked(project_id, db)
    activity = Activity(project_id=project_id, updated_by=current_user.id, **payload.model_dump())
    db.add(activity)
    # Register the physical unit as a side effect (planner types name + location,
    # exactly the spreadsheet habit) — new units start as placeholder slots.
    if payload.rig_name:
        await ensure_registered(
            db, project_id, kind="rig", name=payload.rig_name,
            location=payload.location, user_id=current_user.id,
        )
    elif payload.hwu_name:
        await ensure_registered(
            db, project_id, kind="hwu", name=payload.hwu_name,
            location=None, user_id=current_user.id,
        )
    await db.commit()
    await db.refresh(activity)
    return ActivityResponse.model_validate(activity)


@router.patch("/{activity_id}", response_model=ActivityResponse)
async def update_activity(
    project_id: uuid.UUID,
    activity_id: uuid.UUID,
    payload: ActivityUpdate,
    current_user: CurrentUser,
    db: DB,
) -> ActivityResponse:
    await assert_member(project_id, current_user, db, allowed_roles={ProjectRole.planner})
    result = await db.execute(
        select(Activity).where(
            Activity.id == activity_id, Activity.project_id == project_id
        )
    )
    activity = result.scalar_one_or_none()
    if activity is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Activity not found")

    ensure_activity_unlocked(activity)

    # ── Optimistic lock check ──────────────────────────────────────────────────
    if payload.expected_updated_at is not None:
        db_ts = _normalize_ts(activity.updated_at)
        client_ts = _normalize_ts(payload.expected_updated_at)
        if abs((db_ts - client_ts).total_seconds()) > 1:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={
                    "code": "conflict",
                    "message": "Another user modified this activity after you loaded it.",
                    "updated_by": activity.updated_by_name or "Unknown",
                    "updated_at": activity.updated_at.isoformat(),
                },
            )

    # ── Apply changes & write audit log ───────────────────────────────────────
    changes = payload.model_dump(exclude_unset=True, exclude=_AUDIT_EXCLUDE)
    # An activity is scheduled on a rig OR an HWU, never both. Check the MERGED
    # state — a field absent from this payload keeps its current value.
    if changes.get("rig_name", activity.rig_name) and changes.get(
        "hwu_name", activity.hwu_name
    ):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="an activity uses either a rig or an HWU, not both; clear one first",
        )
    for field, new_val in changes.items():
        old_val = getattr(activity, field)
        if old_val != new_val:
            db.add(AuditLog(
                project_id=project_id,
                user_id=current_user.id,
                entity_type="activity",
                entity_id=activity_id,
                field=field,
                old_value=str(old_val) if old_val is not None else None,
                new_value=str(new_val) if new_val is not None else None,
            ))
        setattr(activity, field, new_val)

    activity.updated_by = current_user.id
    activity.updated_at = datetime.now(timezone.utc)  # explicit for microsecond precision
    # Register the (possibly re-pointed) unit — no-op when it already exists.
    if activity.rig_name:
        await ensure_registered(
            db, project_id, kind="rig", name=activity.rig_name,
            location=activity.location, user_id=current_user.id,
        )
    elif activity.hwu_name:
        await ensure_registered(
            db, project_id, kind="hwu", name=activity.hwu_name,
            location=None, user_id=current_user.id,
        )
    await db.commit()
    await db.refresh(activity)
    return ActivityResponse.model_validate(activity)


async def _set_completion(
    project_id: uuid.UUID,
    activity_id: uuid.UUID,
    completed: bool,
    current_user: User,
    db: AsyncSession,
) -> ActivityResponse:
    await assert_member(project_id, current_user, db, allowed_roles={ProjectRole.planner})
    result = await db.execute(
        select(Activity).where(
            Activity.id == activity_id, Activity.project_id == project_id
        )
    )
    activity = result.scalar_one_or_none()
    if activity is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Activity not found")

    ensure_activity_unlocked(activity)

    new_value = datetime.now(timezone.utc) if completed else None
    db.add(AuditLog(
        project_id=project_id,
        user_id=current_user.id,
        entity_type="activity",
        entity_id=activity_id,
        field="completed_at",
        old_value=activity.completed_at.isoformat() if activity.completed_at else None,
        new_value=new_value.isoformat() if new_value else None,
    ))
    activity.completed_at = new_value
    activity.updated_by = current_user.id
    activity.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(activity)
    return ActivityResponse.model_validate(activity)


@router.post("/{activity_id}/complete", response_model=ActivityResponse)
async def complete_activity(
    project_id: uuid.UUID, activity_id: uuid.UUID, current_user: CurrentUser, db: DB
) -> ActivityResponse:
    """Close a completed activity. Completed activities are dropped when the
    project is cloned into the next quarter."""
    return await _set_completion(project_id, activity_id, True, current_user, db)


@router.post("/{activity_id}/reopen", response_model=ActivityResponse)
async def reopen_activity(
    project_id: uuid.UUID, activity_id: uuid.UUID, current_user: CurrentUser, db: DB
) -> ActivityResponse:
    """Reopen a previously completed activity."""
    return await _set_completion(project_id, activity_id, False, current_user, db)


@router.delete("/{activity_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_activity(
    project_id: uuid.UUID, activity_id: uuid.UUID, current_user: CurrentUser, db: DB
) -> None:
    await assert_member(project_id, current_user, db, allowed_roles={ProjectRole.planner})
    result = await db.execute(
        select(Activity).where(
            Activity.id == activity_id, Activity.project_id == project_id
        )
    )
    activity = result.scalar_one_or_none()
    if activity is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Activity not found")

    ensure_activity_unlocked(activity)
    await db.delete(activity)
    await db.commit()


@router.get("/{activity_id}/history", response_model=list[AuditEntryResponse])
async def get_activity_history(
    project_id: uuid.UUID,
    activity_id: uuid.UUID,
    current_user: CurrentUser,
    db: DB,
    limit: int = Query(default=50, le=200),
) -> list[AuditEntryResponse]:
    # Reads are org-wide: any authenticated user may view campaign data.
    result = await db.execute(
        select(AuditLog)
        .where(
            AuditLog.entity_type == "activity",
            AuditLog.entity_id == activity_id,
            AuditLog.project_id == project_id,
        )
        .order_by(AuditLog.timestamp.desc())
        .limit(limit)
    )
    return [AuditEntryResponse.model_validate(e) for e in result.scalars().all()]




# ── Activity-type resolution at import (see services/data_processor.py) ───────
# Layer order: canonical → formatting (silent) → curated alias → this upload's
# manual mapping → unknown (imported verbatim, warned). Word-level rewrites are
# reported in the response; the dry run feeds the mapping dialog.

_MAX_TYPE_MAPPINGS = 50


def _parse_type_mappings(
    mappings_raw: str | None, remember_raw: str | None
) -> tuple[dict[str, str], dict[str, str], list[str]]:
    """Validate the mapping form fields (allow-list: targets must be canonical).

    Returns (mappings_by_key, display_by_key, remember_keys): resolution wants
    normalized keys; the remember step wants the sheet's original wording too.
    """
    if not mappings_raw:
        return {}, {}, []
    try:
        parsed = json.loads(mappings_raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="mappings must be a JSON object")
    if not isinstance(parsed, dict) or len(parsed) > _MAX_TYPE_MAPPINGS:
        raise HTTPException(
            status_code=422,
            detail=f"mappings must be a JSON object with at most {_MAX_TYPE_MAPPINGS} entries",
        )
    by_key: dict[str, str] = {}
    display_by_key: dict[str, str] = {}
    for source, target in parsed.items():
        if not isinstance(source, str) or not isinstance(target, str):
            raise HTTPException(status_code=422, detail="mappings entries must be strings")
        source = source.strip()
        if not source or len(source) > 128:
            raise HTTPException(status_code=422, detail="mapping sources must be 1-128 characters")
        if target not in CANONICAL_ACTIVITY_TYPES:
            raise HTTPException(
                status_code=422,
                detail=f"'{target}' is not a canonical activity type",
            )
        key = normalize_activity_type_key(source)
        # Bound the NORMALIZED key, not just the raw source: normalization can
        # lengthen a value (it inserts a space before every "("), and this key is
        # stored in activity_type_aliases.alias_key (String(128)).
        if len(key) > 128:
            raise HTTPException(
                status_code=422, detail="mapping source is too long after normalisation"
            )
        # Two sheet wordings that normalise to the same unit must not map to
        # different types — the stored alias is keyed by the normalized form, so
        # silently letting the last one win would drop a mapping.
        if key in by_key and by_key[key] != target:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"'{source}' conflicts with another mapping for the same "
                    f"normalised value — they can't map to different types."
                ),
            )
        by_key[key] = target
        display_by_key[key] = source

    remember_keys: list[str] = []
    if remember_raw:
        try:
            remember_list = json.loads(remember_raw)
        except json.JSONDecodeError:
            raise HTTPException(status_code=422, detail="remember must be a JSON array")
        # Explicit bound, like `mappings` — a remember key must be one of the
        # (already-capped) mapping sources, so it can never exceed that count.
        if not isinstance(remember_list, list) or len(remember_list) > _MAX_TYPE_MAPPINGS:
            raise HTTPException(
                status_code=422,
                detail=f"remember must be a JSON array with at most {_MAX_TYPE_MAPPINGS} entries",
            )
        for source in remember_list:
            key = normalize_activity_type_key(str(source))
            if key not in by_key:
                raise HTTPException(
                    status_code=422,
                    detail="remember entries must be sources present in mappings",
                )
            remember_keys.append(key)
    return by_key, display_by_key, remember_keys


async def _load_type_aliases(db: AsyncSession) -> dict[str, str]:
    rows = (await db.execute(select(ActivityTypeAlias))).scalars().all()
    return {row.alias_key: row.canonical for row in rows}


# Reporting accumulators for the mapping dialog + import summary. `unknown` is
# keyed by NORMALIZED key so whitespace/case variants of one value collapse to a
# single dialog entry (keeping the first-seen display) — matching the key space
# _parse_type_mappings maps them in, so a variant pair can't split a mapping.
_TypeTally = tuple[dict[str, tuple[str, int]], dict[tuple[str, str], int]]


def _resolve_type_in_place(
    fields: dict, db_aliases: dict[str, str], user_mappings: dict[str, str]
) -> tuple[str | None, str, str]:
    """Resolve fields['activity_type'] IN PLACE; return (raw, resolved, how) so
    the caller decides whether to tally it (the long path tallies only rows that
    survive validation — see #_tally_type)."""
    raw = fields.get("activity_type")
    resolved, how = resolve_activity_type(raw, db_aliases, user_mappings)
    if how in ("formatting", "alias", "mapped"):
        fields["activity_type"] = resolved
    return raw, resolved, how


def _tally_type(
    raw: str | None, resolved: str | None, how: str, tally: _TypeTally
) -> None:
    """Record one resolved row for reporting. Word-level rewrites (alias/mapped)
    are reported; formatting fixes stay silent (same words). Unknowns dedupe by
    normalized key."""
    unknown, applied = tally
    if how in ("alias", "mapped") and raw and resolved:
        pair = (raw.strip(), resolved)
        applied[pair] = applied.get(pair, 0) + 1
    elif how == "unknown" and raw and str(raw).strip():
        display = str(raw).strip()
        key = normalize_activity_type_key(display)
        seen_display, count = unknown.get(key, (display, 0))
        unknown[key] = (seen_display, count + 1)


def _preview_payload(tally: _TypeTally) -> dict:
    unknown, applied = tally
    return {
        "unknown_types": [
            {"value": display, "rows": rows}
            for _key, (display, rows) in sorted(unknown.items())
        ],
        "applied_mappings": [
            {"source": source, "target": target, "rows": rows}
            for (source, target), rows in sorted(applied.items())
        ],
    }


async def _remember_type_aliases(
    db: AsyncSession,
    remember_keys: list[str],
    mappings_by_key: dict[str, str],
    display_by_key: dict[str, str],
    current_user: User,
) -> None:
    """Persist "remember this mapping" choices so future uploads resolve them
    automatically. Added to the import's session — committed atomically with
    the import itself; each is a GLOBAL governance event (it changes how every
    future upload is read)."""
    for key in dict.fromkeys(remember_keys):  # dedupe, keep order
        canonical = mappings_by_key[key]
        display = display_by_key[key]
        existing = (
            await db.execute(
                select(ActivityTypeAlias).where(ActivityTypeAlias.alias_key == key)
            )
        ).scalar_one_or_none()

        if existing is None:
            alias = ActivityTypeAlias(
                alias_key=key,
                alias_display=display,
                canonical=canonical,
                created_by=current_user.id,
            )
            try:
                # Savepoint so a lost race on the unique alias_key doesn't abort
                # the whole import — settle it the way services/registry.py's
                # ensure_registered does for its own unique-identity insert.
                async with db.begin_nested():
                    db.add(alias)
                    await db.flush()
            except IntegrityError:
                # A concurrent import created it first — fall through and update.
                existing = (
                    await db.execute(
                        select(ActivityTypeAlias).where(ActivityTypeAlias.alias_key == key)
                    )
                ).scalar_one_or_none()
            else:
                db.add(
                    governance_event(
                        project_id=None,
                        user_id=current_user.id,
                        entity_type=ENTITY_VOCABULARY,
                        entity_id=alias.id,
                        action="activity_type_alias_added",
                        detail=f"Import mapping remembered: '{display}' → '{canonical}'",
                    )
                )
                continue

        # Existing row (pre-found, or won by a concurrent import): only touch it
        # when the target actually changed.
        if existing is None or existing.canonical == canonical:
            continue
        existing.canonical = canonical
        existing.alias_display = display
        db.add(
            governance_event(
                project_id=None,
                user_id=current_user.id,
                entity_type=ENTITY_VOCABULARY,
                entity_id=existing.id,
                action="activity_type_alias_updated",
                detail=f"Import mapping remembered: '{display}' → '{canonical}'",
            )
        )


@router.post("/import", response_model=ImportResponse)
async def import_activities(
    project_id: uuid.UUID,
    current_user: CurrentUser,
    db: DB,
    file: UploadFile = File(...),
    replace: bool = Query(default=True, description="Replace all existing activities"),
    dry_run: bool = Query(
        default=False,
        description="Preview: validate and resolve activity types, write NOTHING; "
        "returns unknown_types for the mapping dialog.",
    ),
    # Multipart form fields (they ride alongside the file):
    # mappings — JSON object {sheet value → canonical type} from the mapping step;
    # remember — JSON array of mapping sources to persist as aliases.
    mappings: str | None = Form(default=None),
    remember: str | None = Form(default=None),
) -> ImportResponse:
    """Upload a CSV or Excel file and bulk-insert activities into the project.

    Every row is validated through the same `ActivityCreate` schema the JSON API
    uses, so an import cannot smuggle in NULL/invalid dates, end-before-start
    ranges, or non-canonical enum values that a direct POST would reject. The file
    is validated in full *before* any write, so a single bad row never deletes the
    existing schedule (replace mode) or leaves a partial import behind.
    """
    await assert_member(project_id, current_user, db, allowed_roles={ProjectRole.planner})

    # A bulk import would delete/replace activities that may be frozen under a
    # pending revision — refuse while any are locked.
    await assert_project_not_locked(project_id, db)

    mappings_by_key, display_by_key, remember_keys = _parse_type_mappings(mappings, remember)
    db_aliases = await _load_type_aliases(db)

    # Cap what we materialize BEFORE parsing (read one byte past the limit so
    # an at-limit file passes and an over-limit one is detected without
    # buffering the rest). Real schedules are well under 1 MB; the cap exists
    # so a huge or decompression-bomb upload can't exhaust server memory.
    content = await file.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"File is larger than the {MAX_IMPORT_BYTES // (1024 * 1024)} MB "
                f"import limit. Split the schedule or remove unrelated sheets."
            ),
        )
    filename = file.filename or ""

    # keep_default_na=False + na_values=[""]: only genuinely EMPTY cells read as
    # missing. Pandas' default NA list would silently turn literal cell text like
    # "N/A" (or a well named "NA") into NaN instead of keeping the planner's value.
    _na = {"keep_default_na": False, "na_values": [""]}
    try:
        if filename.endswith((".xlsx", ".xls")):
            sheets = pd.read_excel(io.BytesIO(content), sheet_name=None, **_na)
            df = sheets.get(SCHEDULE_SHEET)
            if df is None:  # unnamed / legacy single-sheet workbook → first sheet
                df = next(iter(sheets.values())) if sheets else pd.DataFrame()
            # Any other tab (Guidance, a legacy "Readiness" tab, planner scratch
            # sheets) is ignored — readiness is managed in the app, not uploaded.
        else:
            df = pd.read_csv(io.BytesIO(content), **_na)
    except Exception as exc:
        # Don't echo the parser's internal message (it can include file content).
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Could not parse the uploaded file. Provide a valid CSV or Excel file.",
        ) from exc

    # Row ceiling: the byte cap alone can't bound a compressed .xlsx, which can
    # expand far past its file size when parsed.
    if len(df) > MAX_IMPORT_ROWS:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"The sheet has {len(df)} rows — more than the {MAX_IMPORT_ROWS} "
                f"the import accepts. Split the schedule into smaller files."
            ),
        )

    # The schedule workbook is one row per activity (with per-rig contract
    # expiry). It has its own ingestion path; the legacy wide CSV path
    # continues below.
    if is_long_schedule(df):
        return await _import_long_schedule(
            df,
            project_id,
            current_user,
            db,
            replace,
            dry_run=dry_run,
            db_aliases=db_aliases,
            user_mappings=mappings_by_key,
            remember_keys=remember_keys,
            display_by_key=display_by_key,
        )

    try:
        validate_csv_columns(df)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)
        ) from exc

    rows = csv_df_to_db_rows(df, str(project_id))
    # The wide path is all-or-nothing (any invalid row → 422 below, nothing
    # imported), so tallying every row's resolution is faithful to the outcome.
    tally: _TypeTally = ({}, {})
    for row in rows:
        raw, resolved, how = _resolve_type_in_place(row, db_aliases, mappings_by_key)
        _tally_type(raw, resolved, how, tally)

    # Validate every row against the schema before touching the database.
    validated: list[ActivityCreate] = []
    errors: list[str] = []
    for i, row in enumerate(rows):
        fields = {k: v for k, v in row.items() if k != "project_id"}
        try:
            validated.append(ActivityCreate(**fields))
        except ValidationError as exc:
            for err in exc.errors():
                loc = ".".join(str(p) for p in err["loc"]) or "row"
                errors.append(f"Row {i + 2}: {loc} — {err['msg']}")
    if errors:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "message": "Import rejected — fix the rows below and re-upload.",
                "errors": errors[:20],
            },
        )

    if dry_run:
        # Preview: nothing written, nothing deleted, nothing remembered.
        return ImportResponse(
            imported=len(validated),
            replaced=False,
            dry_run=True,
            warnings=unknown_activity_type_warnings([m.activity_type for m in validated])[:200],
            **_preview_payload(tally),
        )

    if replace:
        await db.execute(delete(Activity).where(Activity.project_id == project_id))

    for model in validated:
        db.add(
            Activity(
                project_id=project_id,
                updated_by=current_user.id,
                **model.model_dump(),
            )
        )

    await ensure_activity_resources_registered(
        db, project_id, _resource_triples(validated), current_user.id
    )
    await _remember_type_aliases(db, remember_keys, mappings_by_key, display_by_key, current_user)
    await db.commit()
    return ImportResponse(
        imported=len(validated),
        replaced=replace,
        warnings=unknown_activity_type_warnings([m.activity_type for m in validated])[:200],
        **_preview_payload(tally),
    )


def _resource_triples(models: list[ActivityCreate]) -> set[tuple[str, str, str | None]]:
    """Distinct physical units referenced by a batch of activities — rigs keyed
    with their terrain, HWUs without (they are mobile; see services/registry.py)."""
    triples: set[tuple[str, str, str | None]] = set()
    for m in models:
        if m.rig_name:
            triples.add(("rig", m.rig_name, m.location))
        elif m.hwu_name:
            triples.add(("hwu", m.hwu_name, None))
    return triples


async def _import_long_schedule(
    df: pd.DataFrame,
    project_id: uuid.UUID,
    current_user: User,
    db: AsyncSession,
    replace: bool,
    *,
    dry_run: bool = False,
    db_aliases: dict[str, str] | None = None,
    user_mappings: dict[str, str] | None = None,
    remember_keys: list[str] | None = None,
    display_by_key: dict[str, str] | None = None,
) -> ImportResponse:
    """Ingest the schedule workbook: one activity per Schedule row, plus
    per-rig/HWU contract expiry.

    Validated in full before any write, so a bad row never leaves a partial import
    or deletes the existing schedule. Readiness is NOT read from the upload — it
    is per-project state managed in the app, and it survives a replace import
    (the gates describe the field project, not the schedule rows).
    """
    try:
        parsed, rig_contracts, hwu_contracts = parse_long_schedule(df)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)
        ) from exc

    # Validate each activity (same ActivityCreate gate the JSON API uses). An
    # invalid row is skipped (not imported) and reported. Type resolution happens
    # per row and is tallied ONLY after the row validates, so the reported
    # unknown/mapped counts match what was actually imported.
    validated: list[ActivityCreate] = []
    skipped_rows: list[dict[str, str]] = []
    warnings: list[str] = []
    tally: _TypeTally = ({}, {})
    for pa in parsed:
        label = pa.fields.get("well_name") or pa.fields.get("activity_type") or "row"
        raw, resolved, how = _resolve_type_in_place(
            pa.fields, db_aliases or {}, user_mappings or {}
        )
        try:
            activity_in = ActivityCreate(**pa.fields)
        except ValidationError as exc:
            reason = "; ".join(
                f"{'.'.join(str(p) for p in err['loc']) or 'field'} — {err['msg']}"
                for err in exc.errors()
            )
            skipped_rows.append({"well": label, "reason": reason})
            continue
        _tally_type(raw, resolved, how, tally)  # only kept rows count toward the summary
        validated.append(activity_in)

    # Non-canonical activity types import fine but chart in neutral grey — tell
    # the planner so the vocabulary gets fixed (or the catalogue extended).
    warnings.extend(unknown_activity_type_warnings([a.activity_type for a in validated]))
    # A RIG name spanning terrains is legal (two physical units under the
    # planner's naming convention) but can also hide a location typo — warn.
    # HWUs are exempt: they are mobile, so cross-terrain use is unremarkable.
    warnings.extend(
        cross_terrain_resource_warnings([(a.rig_name, a.location) for a in validated], "Rig")
    )

    if dry_run:
        # Preview: nothing written, nothing deleted, nothing remembered.
        return ImportResponse(
            imported=len(validated),
            replaced=False,
            skipped=len(skipped_rows),
            skipped_rows=skipped_rows[:200],
            warnings=warnings[:200],
            dry_run=True,
            **_preview_payload(tally),
        )

    # Replace only when at least one well is valid — never wipe the schedule to
    # import nothing (e.g. an entirely-bad file in replace mode).
    if replace and validated:
        # Activities only — ProjectReadiness rows deliberately SURVIVE a replace:
        # gate statuses describe the field project, not the schedule rows, and a
        # re-uploaded schedule must never silently reset what the planner set in
        # the app. A project that drops out of the schedule just stops being
        # listed (GET /readiness derives its projects from activities) and
        # resurfaces with its statuses intact if the project returns.
        await db.execute(delete(Activity).where(Activity.project_id == project_id))

    for activity_in in validated:
        db.add(
            Activity(
                project_id=project_id,
                updated_by=current_user.id,
                **activity_in.model_dump(),
            )
        )

    # Register every physical unit the sheet references (rigs terrain-qualified,
    # HWUs by name) — new units start as placeholder slots.
    await ensure_activity_resources_registered(
        db, project_id, _resource_triples(validated), current_user.id
    )

    # Resource contract expiry — upsert each rig / HWU end date, so an imported
    # sheet sets the contract exactly like the editor (it drives the
    # contract-expiry marker). Rig contracts are per PHYSICAL unit: keyed
    # (name, terrain), so a LAND and SWAMP rig sharing a name stay separate.
    # Rows are matched like the registry — trimmed, case-insensitively — so a
    # sheet whose casing drifted from an earlier save updates the SAME unit's
    # row instead of splitting it; legacy case-variant duplicates all get the
    # new expiry (kept consistent, never guessed between).
    for (rig_name, terrain), expiry in rig_contracts.items():
        existing_rows = await lane_contract_rows(
            db, project_id, kind="rig", name=rig_name, terrain=terrain
        )
        if existing_rows:
            for existing in existing_rows:
                existing.contract_end = expiry
                existing.updated_by = current_user.id
        else:
            db.add(
                RigContract(
                    project_id=project_id,
                    rig_name=rig_name,
                    terrain=terrain,
                    contract_end=expiry,
                    updated_by=current_user.id,
                )
            )

    for hwu_name, expiry in hwu_contracts.items():
        existing_hwu_rows = await lane_contract_rows(
            db, project_id, kind="hwu", name=hwu_name
        )
        if existing_hwu_rows:
            for existing_hwu in existing_hwu_rows:
                existing_hwu.contract_end = expiry
                existing_hwu.updated_by = current_user.id
        else:
            db.add(
                HwuContract(
                    project_id=project_id,
                    hwu_name=hwu_name,
                    contract_end=expiry,
                    updated_by=current_user.id,
                )
            )

    await _remember_type_aliases(
        db, remember_keys or [], user_mappings or {}, display_by_key or {}, current_user
    )
    await db.commit()
    return ImportResponse(
        imported=len(validated),
        replaced=replace,
        skipped=len(skipped_rows),
        skipped_rows=skipped_rows[:200],
        warnings=warnings[:200],
        **_preview_payload(tally),
    )
