"""Rig fleet optimization endpoints (docs/rig-optimization-spec.md).

Planner-only (assert_can_plan; admins pass implicitly): the tool is a planning
calculator. It reads nothing campaign-scoped and writes nothing — no campaign,
revision, approval, or governance state is touched (spec §8).
"""

import io
import logging
import uuid
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse

from app.config import settings
from app.core.auth import get_current_user
from app.core.rbac import assert_can_plan
from app.models.user import User
from app.schemas.optimizer import (
    MAX_PROJECTS,
    DemandRow,
    OptimizationRequest,
    OptimizationResponse,
    ParsedScheduleResponse,
    Terrain,
    TerrainResultOut,
)
from app.services.rig_optimizer import Assumptions, Options, optimize

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/optimizer", tags=["optimizer"])

CurrentUser = Annotated[User, Depends(get_current_user)]

_MAX_UPLOAD_BYTES = 1 * 1024 * 1024  # schedule tables are tiny; 1 MB is generous


@router.post("/rig-fleet", response_model=OptimizationResponse)
async def optimize_rig_fleet(
    payload: OptimizationRequest, current_user: CurrentUser
) -> OptimizationResponse:
    """Compute the minimum rig fleet per terrain for the given schedule."""
    assert_can_plan(current_user)

    engine = (settings.optimizer_engine or "heuristic").strip().lower()
    warning = None
    if engine == "milp":
        # The exact-solver engine needs a MILP/CP library (e.g. OR-Tools) that
        # has not been adopted — new dependencies go through IT review first
        # (CLAUDE.md). Fall back loudly rather than silently.
        try:
            import ortools  # type: ignore  # noqa: F401
        except ImportError:
            engine = "heuristic"
            warning = (
                "MILP engine is configured but its solver library is not "
                "installed; results computed with the heuristic engine."
            )
            logger.warning("optimizer_engine=milp requested but OR-Tools missing")
        else:  # pragma: no cover — exercised once the dependency is approved
            engine = "heuristic"
            warning = "MILP engine not yet implemented; heuristic engine used."

    results = optimize(
        demand_rows=[
            {
                "terrain": row.terrain.value,
                "project": row.project,
                "wells_by_year": row.wells_by_year,
            }
            for row in payload.demand
        ],
        assumptions=Assumptions(**payload.assumptions.model_dump()),
        options=Options(**payload.options.model_dump()),
    )

    run_id = uuid.uuid4()
    logger.info(
        "rig optimization run id=%s user_id=%s terrains=%d projects=%d engine=%s",
        run_id,
        current_user.id,
        len(results),
        len(payload.demand),
        engine,
    )
    return OptimizationResponse(
        run_id=run_id,
        engine=engine,  # type: ignore[arg-type]
        warning=warning,
        results=[
            TerrainResultOut(
                terrain=Terrain(r.terrain),
                feasible=r.feasible,
                rig_count=r.rig_count,
                rigs=[
                    {
                        "name": rig.name,
                        "wells": [w.__dict__ for w in rig.wells],
                    }
                    for rig in r.rigs
                ],
                rigs_active_per_year=r.rigs_active_per_year,
                utilization_per_rig=r.utilization_per_rig,
                binding=r.binding,
                infeasible_wells=r.infeasible_wells,
            )
            for r in results
        ],
    )


@router.post("/rig-fleet/export")
async def export_rig_fleet(
    payload: OptimizationRequest, current_user: CurrentUser
) -> StreamingResponse:
    """Run the optimization and return the full result as an Excel workbook:
    Summary (rigs per terrain), Rig Schedule (every well with dates and gaps),
    and the input Demand table. Same validation and planner gate as the run."""
    assert_can_plan(current_user)

    results = optimize(
        demand_rows=[
            {
                "terrain": row.terrain.value,
                "project": row.project,
                "wells_by_year": row.wells_by_year,
            }
            for row in payload.demand
        ],
        assumptions=Assumptions(**payload.assumptions.model_dump()),
        options=Options(**payload.options.model_dump()),
    )

    from openpyxl import Workbook  # already a vetted dependency (Excel import)
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)
    date_fmt = "DD-MM-YYYY"  # matches the app-wide date rendition

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Terrain", "Feasible", "Rigs required", "Binding project", "Binding year"])
    for c in ws[1]:
        c.font = bold
    for r in results:
        ws.append(
            [
                r.terrain,
                "Yes" if r.feasible else "No — infeasible",
                r.rig_count if r.feasible else None,
                (r.binding or {}).get("project"),
                (r.binding or {}).get("year"),
            ]
        )
    ws.append([])
    ws.append(["Terrain", "Year", "Rigs active"])
    for c in ws[ws.max_row]:
        c.font = bold
    for r in results:
        for year, n in sorted(r.rigs_active_per_year.items()):
            ws.append([r.terrain, year, n])

    sched = wb.create_sheet("Rig Schedule")
    sched.append(
        ["Terrain", "Rig", "Project", "Well", "Committed year",
         "Start", "End", "Gap before (days)", "Gap type"]
    )
    for c in sched[1]:
        c.font = bold
    gap_labels = {
        "none": "", "inter_well": "Rig move (2 wk)",
        "batch": "Batch gap (4 wk)", "project_move": "Project move (45 d)",
    }
    for r in results:
        for rig in r.rigs:
            for w in rig.wells:
                sched.append(
                    [r.terrain, rig.name, w.project, w.label.split("·")[-1].strip(),
                     w.year, w.start, w.end, w.gap_before_days or None,
                     gap_labels.get(w.gap_kind, w.gap_kind)]
                )
                sched.cell(row=sched.max_row, column=6).number_format = date_fmt
                sched.cell(row=sched.max_row, column=7).number_format = date_fmt

    demand_ws = wb.create_sheet("Demand")
    years = sorted({int(y) for row in payload.demand for y in row.wells_by_year})
    demand_ws.append(["Terrain", "Project", *[str(y) for y in years]])
    for c in demand_ws[1]:
        c.font = bold
    for row in payload.demand:
        demand_ws.append(
            [row.terrain.value, row.project, *[row.wells_by_year.get(y) for y in years]]
        )

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(
        "rig optimization export user_id=%s terrains=%d", current_user.id, len(results)
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="rig-optimization.xlsx"'},
    )


@router.post("/parse-schedule", response_model=ParsedScheduleResponse)
async def parse_schedule(file: UploadFile, current_user: CurrentUser) -> ParsedScheduleResponse:
    """Parse an uploaded wells schedule (CSV or Excel) into demand rows.

    Expected layout: `Terrain | Project | <year> | <year> | …` — exactly the
    planning table format. Unparseable rows are reported in `issues`, never
    silently dropped.
    """
    assert_can_plan(current_user)

    raw = await file.read()
    if len(raw) > _MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Schedule file too large (limit 1 MB)",
        )

    import pandas as pd  # already a vetted dependency (CSV import feature)

    name = (file.filename or "").lower()
    try:
        if name.endswith((".xlsx", ".xlsm")):
            df = pd.read_excel(io.BytesIO(raw))
        else:
            df = pd.read_csv(io.BytesIO(raw))
    except Exception:
        logger.warning("schedule parse failed filename=%r", file.filename)
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Could not read the file — expected a CSV or Excel table",
        )

    cols = [str(c).strip() for c in df.columns]
    lower = [c.lower() for c in cols]
    if "terrain" not in lower or "project" not in lower:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Missing required columns: Terrain, Project",
        )
    terrain_col = cols[lower.index("terrain")]
    project_col = cols[lower.index("project")]
    year_cols: dict[str, int] = {}
    for c in cols:
        try:
            year = int(float(c))
        except ValueError:
            continue
        if 2000 <= year <= 2100:
            year_cols[c] = year
    if not year_cols:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="No year columns found (expected e.g. 2027, 2028, …)",
        )

    terrain_alias = {t.value.lower(): t for t in Terrain}
    # Common spellings for shallow water offshore — all read as SWO.
    terrain_alias["offshore"] = Terrain.swo
    terrain_alias["shallow offshore"] = Terrain.swo
    demand: list[DemandRow] = []
    issues: list[str] = []
    for idx, row in df.iterrows():
        rownum = int(idx) + 2  # 1-based + header, matching what the user sees
        terrain_raw = str(row.get(terrain_col, "")).strip()
        project = str(row.get(project_col, "")).strip()
        if not terrain_raw or terrain_raw.lower() == "nan":
            continue  # blank spacer row
        terrain = terrain_alias.get(terrain_raw.lower())
        if terrain is None:
            issues.append(
                f"Row {rownum}: unknown terrain '{terrain_raw}' (expected Land, Swamp, or SWO)"
            )
            continue
        if not project or project.lower() == "nan":
            issues.append(f"Row {rownum}: missing project name")
            continue
        wells_by_year: dict[int, int] = {}
        bad_cell = False
        for col, year in year_cols.items():
            value = row.get(col)
            if pd.isna(value) or str(value).strip() == "":
                continue
            try:
                wells = int(float(value))
            except (TypeError, ValueError):
                issues.append(f"Row {rownum}: '{value}' in {year} is not a number")
                bad_cell = True
                break
            if wells < 0:
                issues.append(f"Row {rownum}: negative well count in {year}")
                bad_cell = True
                break
            if wells > 0:
                wells_by_year[year] = wells
        if bad_cell:
            continue
        if not wells_by_year:
            issues.append(f"Row {rownum}: {project} has no wells in any year")
            continue
        demand.append(
            DemandRow(terrain=terrain, project=project, wells_by_year=wells_by_year)
        )
        if len(demand) > MAX_PROJECTS:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Too many project rows (limit {MAX_PROJECTS})",
            )

    return ParsedScheduleResponse(
        demand=demand, years=sorted(set(year_cols.values())), issues=issues
    )
